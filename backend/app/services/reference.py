from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from functools import lru_cache
from pathlib import Path
from typing import Any

from app.core.config import get_settings


CURATED_DATASET_CANDIDATE_NAMES = (
    "campus_jobs_augmented.csv",
    "campus_jobs_curated.csv",
    "a13_jobs_augmented.csv",
    "a13_jobs_supplement.csv",
)

DATASET_CANDIDATE_NAMES = (
    *CURATED_DATASET_CANDIDATE_NAMES,
    "official_jobs.xls",
    "official_jobs.xlsx",
    "official_jobs.csv",
    "a13_jobs.xls",
    "a13_jobs.xlsx",
    "a13_jobs.csv",
    "sample_jobs.csv",
)

JOB_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "title": ("title", "岗位名称", "职位名称", "岗位名", "职位名"),
    "location": ("location", "地址", "工作地址", "工作地点", "城市"),
    "salary_range": ("salary_range", "薪资范围", "薪资", "薪酬范围", "工资范围"),
    "company_name": ("company_name", "公司名称", "公司全称", "企业名称"),
    "industry": ("industry", "所属行业", "行业"),
    "company_size": ("company_size", "公司规模", "人员规模", "企业规模"),
    "ownership_type": ("ownership_type", "公司类型", "企业性质", "公司性质"),
    "job_code": ("job_code", "岗位编码", "职位编码", "岗位id", "职位id"),
    "description": ("description", "岗位详情", "职位描述", "岗位描述", "工作内容"),
    "company_intro": ("company_intro", "公司详情", "公司简介", "企业简介"),
    "source_url": ("source_url", "岗位来源地址", "来源地址", "职位链接", "岗位链接"),
    "updated_at": ("updated_at", "更新日期", "发布时间"),
}

COMPUTER_INDUSTRY_KEYWORDS = (
    "互联网",
    "计算机",
    "软件",
    "信息技术",
    "it服务",
    "通信",
    "网络设备",
    "电子设备",
    "人工智能",
    "ai",
    "大数据",
    "云计算",
    "物联网",
    "半导体",
    "集成电路",
    "智能制造",
    "信息化",
)

COMPUTER_ROLE_KEYWORDS = (
    "开发",
    "软件",
    "前端",
    "后端",
    "全栈",
    "java",
    "python",
    "golang",
    "go开发",
    "c++",
    "c#",
    "php",
    "测试",
    "算法",
    "数据",
    "ai",
    "人工智能",
    "机器学习",
    "深度学习",
    "运维",
    "devops",
    "架构师",
    "网络",
    "安全",
    "信息安全",
    "嵌入式",
    "硬件",
    "fpga",
    "芯片",
    "通信",
    "实施",
    "解决方案",
    "技术支持",
    "系统工程师",
    "产品经理",
    "项目经理",
    "ui",
    "ue",
)

COMPUTER_SKILL_KEYWORDS = (
    "linux",
    "windows server",
    "数据库",
    "sql",
    "mysql",
    "postgresql",
    "oracle",
    "redis",
    "docker",
    "kubernetes",
    "spring",
    "react",
    "vue",
    "node",
    "html",
    "css",
    "javascript",
    "typescript",
    "java",
    "python",
    "c++",
    "c#",
    "go",
    "算法",
    "模型",
    "深度学习",
    "机器学习",
    "图像处理",
    "网络协议",
    "tcp/ip",
    "自动化测试",
    "云平台",
    "信息化系统",
    "erp",
    "mes",
    "plc",
    "cad",
    "solidworks",
)

NON_COMPUTER_ROLE_KEYWORDS = (
    "人事",
    "行政",
    "财务",
    "会计",
    "出纳",
    "法务",
    "销售",
    "文员",
    "客服",
    "仓管",
    "采购",
)

TITLE_CLEAN_PATTERNS = (
    r"[【\[].*?[】\]]",
    r"[（(].*?[)）]",
    r"\b\d{2}届\b",
    r"\b\d+[kK]\b",
    r"\b\d+薪\b",
)

TITLE_NOISE_KEYWORDS = (
    "急聘",
    "高薪",
    "双休",
    "五险一金",
    "五险",
    "包住",
    "包吃",
    "可小白",
    "接受小白",
    "接受应届",
    "应届生",
    "实习生",
    "周末双休",
    "带薪培训",
    "面试就过",
    "班车",
)

TITLE_CANONICAL_RULES: list[tuple[tuple[str, ...], str]] = [
    (("web前端",), "前端开发"),
    (("前端", "开发"), "前端开发"),
    (("java",), "Java"),
    (("c/c++",), "C/C++"),
    (("c++",), "C/C++"),
    (("软件", "测试"), "软件测试"),
    (("测试开发",), "测试工程师"),
    (("测试", "工程师"), "测试工程师"),
    (("硬件", "测试"), "硬件测试"),
    (("芯片", "测试"), "硬件测试"),
    (("射频", "测试"), "硬件测试"),
    (("质量管理",), "质量管理/测试"),
    (("质检",), "质检员"),
    (("技术支持",), "技术支持工程师"),
    (("售前", "技术支持"), "技术支持工程师"),
    (("售后", "技术支持"), "技术支持工程师"),
    (("fae",), "技术支持工程师"),
    (("实施", "工程师"), "实施工程师"),
    (("项目实施",), "实施工程师"),
    (("运维", "工程师"), "运维工程师"),
    (("应用运维",), "运维工程师"),
    (("算法", "工程师"), "算法工程师"),
    (("人工智能",), "算法工程师"),
    (("数据分析",), "数据分析"),
    (("产品", "专员"), "产品专员/助理"),
    (("产品", "助理"), "产品专员/助理"),
    (("产品", "经理"), "产品经理"),
    (("项目", "专员"), "项目专员/助理"),
    (("项目", "助理"), "项目专员/助理"),
    (("项目", "招投标"), "项目招投标"),
    (("项目", "经理"), "项目经理/主管"),
    (("安全", "工程师"), "安全工程师"),
    (("网络安全",), "安全工程师"),
    (("嵌入式",), "嵌入式开发"),
    (("通信", "工程师"), "通信工程师"),
    (("软件", "开发"), "软件开发工程师"),
    (("后端", "开发"), "后端开发"),
]

TARGET_ROLE_KEYWORDS = (
    "开发",
    "java",
    "c/c++",
    "前端",
    "后端",
    "全栈",
    "测试",
    "硬件",
    "技术支持",
    "实施",
    "运维",
    "算法",
    "数据",
    "产品",
    "项目",
    "安全",
    "网络",
    "嵌入式",
    "通信",
    "软件",
)

STRICT_EXCLUDE_TITLE_KEYWORDS = (
    "客服",
    "销售",
    "广告",
    "运营",
    "推广",
    "bd",
    "大客户",
    "猎头",
    "招聘",
    "培训",
    "翻译",
    "律师",
    "法务",
    "知识产权",
    "专利",
    "档案",
    "资料",
    "储备",
    "总助",
    "商务",
    "审核",
    "顾问",
    "统计员",
)

TECH_EXCEPTION_KEYWORDS = (
    "技术支持",
    "售前",
    "实施",
    "运维",
    "产品",
    "项目",
    "网络安全",
    "安全",
    "数据",
    "算法",
    "开发",
    "测试",
)

CAMPUS_ROLE_KEYWORDS = (
    "开发", "工程师", "分析师", "顾问", "专员", "助理", "管培生", "培训生", "实习",
    "产品", "项目", "运营", "策划", "设计", "编辑", "文案", "销售", "商务", "客户",
    "财务", "会计", "审计", "税务", "法务", "合规", "知识产权", "人力", "招聘",
    "行政", "教师", "讲师", "教研", "咨询", "外贸", "采购", "供应链", "物流",
    "机械", "电气", "土木", "建筑", "质量", "工业", "结构", "临床", "医药",
    "生物", "医学", "营养", "心理", "康复", "食品", "农业", "园艺", "畜牧",
    "环境", "化工", "新能源", "储能", "能源", "石油", "导游", "酒店", "餐饮",
    "体育", "物业", "社区", "社会工作", "翻译", "同声传译",
)

CAMPUS_INDUSTRY_KEYWORDS = (
    "互联网", "计算机", "软件", "信息技术", "人工智能", "大数据", "金融", "银行",
    "证券", "保险", "会计", "法律", "知识产权", "教育", "科研", "传媒", "广告",
    "品牌", "贸易", "人力资源", "企业服务", "制造", "机械", "建筑", "航空",
    "医药", "医疗", "生命科学", "物流", "供应链", "文旅", "酒店", "餐饮",
    "体育", "农业", "食品", "环保", "化工", "新能源", "电力", "设计", "文化",
    "公共服务", "社区", "物业",
)

CAMPUS_EXCLUDE_TITLE_KEYWORDS = (
    "普工", "操作工", "分拣员", "搬运工", "保安", "保洁", "服务员", "收银员",
    "骑手", "司机", "主播", "演员", "模特", "美容师", "美甲", "洗碗工",
)

STUDENT_FACING_INCLUDE_KEYWORDS = (
    "前端开发",
    "后端开发",
    "全栈",
    "测试工程",
    "测试开发",
    "数据分析",
    "数据工程师",
    "算法工程师",
    "ai 算法",
    "人工智能算法",
    "运维工程",
    "devops",
    "sre",
    "网络安全",
    "安全工程师",
    "网络工程师",
    "java",
    "python",
    "c++",
    "c#",
    "go开发",
    "嵌入式",
    "通信工程师",
    "硬件工程师",
    "硬件测试",
    "芯片",
    "fpga",
    "技术支持工程师",
    "实施工程师",
    "解决方案工程师",
    "软件开发",
    "软件工程师",
)

STUDENT_FACING_EXCLUDE_KEYWORDS = (
    "产品经理",
    "产品专员",
    "产品助理",
    "ui/ux",
    "ux",
    "ue",
    "设计师",
    "设计",
    "视觉",
    "交互",
    "运营",
    "市场",
    "广告",
    "销售",
    "财务",
    "会计",
    "法务",
    "审计",
    "培训",
    "讲师",
    "人力",
    "行政",
    "客服",
    "采购",
    "供应链",
    "翻译",
    "职业规划",
    "scrum",
    "敏捷",
    "咨询",
    "顾问",
    "专员",
    "电气",
    "食品",
    "仓储",
    "物流",
    "风控",
    "风险控制",
    "质量管理",
    "招投标",
    "招标",
    "投标",
    "选址",
    "门店",
)

STUDENT_FACING_SKILL_KEYWORDS = (
    "python",
    "java",
    "javascript",
    "typescript",
    "react",
    "vue",
    "next.js",
    "node",
    "fastapi",
    "spring",
    "sql",
    "mysql",
    "postgresql",
    "redis",
    "docker",
    "kubernetes",
    "linux",
    "算法",
    "机器学习",
    "深度学习",
    "pytorch",
    "spark",
    "etl",
    "airflow",
)

INDUSTRY_ALIAS_MAP: dict[str, str] = {
    "it服务": "IT服务",
    "互联网": "互联网",
    "计算机软件": "计算机软件",
    "计算机硬件": "计算机硬件",
    "云计算/大数据": "云计算/大数据",
    "人工智能": "人工智能",
    "网络/信息安全": "网络/信息安全",
    "物联网": "物联网",
    "电子/半导体/集成电路": "电子/半导体/集成电路",
    "电子设备制造": "电子设备制造",
    "通信/网络设备": "通信/网络设备",
    "专业技术服务": "专业技术服务",
    "企业服务": "企业服务",
    "咨询服务": "咨询服务",
    "仪器仪表制造": "仪器仪表制造",
    "电气机械/电力设备": "电气机械/电力设备",
    "电力/水利/热力/燃气": "电力/水利/热力/燃气",
}

OWNERSHIP_ALIAS_MAP: dict[str, str] = {
    "股份制企业": "股份制企业",
    "民营": "民营",
    "民营公司": "民营",
    "上市公司": "上市公司",
    "已上市": "上市公司",
    "国企": "国企",
    "国有企业": "国企",
    "外商独资": "外商独资",
    "合资": "合资",
    "未融资": "未融资",
    "天使轮": "天使轮",
    "a轮": "A轮",
    "b轮": "B轮",
    "c轮": "C轮",
}


INDUSTRY_GROUP_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("AI/大数据", ("人工智能", "云计算", "大数据", "数据服务", "数据智能")),
    ("硬件/半导体", ("半导体", "集成电路", "电子", "计算机硬件", "仪器仪表制造")),
    ("通信/网络", ("通信", "网络设备", "网络/信息安全", "信息安全", "物联网")),
    ("金融科技/数据", ("银行", "保险", "证券", "基金", "信托", "期货")),
    ("医疗信息化", ("医疗设备", "器械", "医药制造", "卫生服务", "生物工程")),
    ("制造业数字化", ("汽车研发", "制造", "工业自动化", "电气机械", "电力设备", "船舶", "航空", "航天", "火车制造")),
    ("能源/电力数字化", ("电力", "水利", "热力", "燃气", "环保", "矿产", "采掘")),
    ("教育/科研信息化", ("学术", "科研", "在线教育")),
    ("软件/互联网", ("计算机软件", "互联网", "IT服务", "电子商务")),
    ("企业服务/咨询", ("专业技术服务", "企业服务", "咨询服务", "人力资源服务")),
]


def _load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def _collapse_text(value: Any) -> str:
    return re.sub(r"\s+", " ", _normalize_text(value)).lower()


def _split_multi_value_text(value: Any) -> list[str]:
    text = _normalize_text(value)
    if not text:
        return []
    text = text.replace("，", ",").replace("、", ",").replace("；", ",").replace("|", ",")
    return [part.strip() for part in re.split(r"[,]+", text) if part and part.strip()]


def _unique_ordered(parts: list[str]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for part in parts:
        key = part.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(part)
    return ordered


def _normalize_location_value(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", "", text.replace("—", "-").replace("--", "-"))


def _normalize_company_size_value(value: Any) -> str:
    parts = [part for part in _split_multi_value_text(value) if not re.fullmatch(r"\d{6,}", part)]
    if not parts:
        return _normalize_text(value)
    return _unique_ordered(parts)[0]


def _normalize_ownership_type_value(value: Any) -> str:
    parts = [part for part in _split_multi_value_text(value) if not re.fullmatch(r"\d{6,}", part)]
    if not parts:
        return _normalize_text(value)
    normalized = [OWNERSHIP_ALIAS_MAP.get(part.lower(), part) for part in parts]
    return _unique_ordered(normalized)[0]


def normalize_industry_value(value: Any) -> str:
    normalized: list[str] = []
    for part in _split_multi_value_text(value):
        cleaned = re.sub(r"\s+", "", part)
        if not cleaned or re.fullmatch(r"\d{6,}", cleaned):
            continue
        normalized.append(INDUSTRY_ALIAS_MAP.get(cleaned.lower(), cleaned))
    if not normalized:
        return _normalize_text(value)
    return " / ".join(_unique_ordered(normalized)[:4])


def split_industry_tags(value: Any) -> list[str]:
    text = normalize_industry_value(value)
    if not text:
        return []
    return _unique_ordered(
        [part.strip() for part in re.split(r"\s*/\s*", text) if part and part.strip()]
    )


def summarize_industry_tags(value: Any, limit: int = 3) -> str:
    tags = split_industry_tags(value)
    if not tags:
        return ""
    return " / ".join(tags[:limit])


def derive_industry_group(value: Any) -> str:
    tags = split_industry_tags(value)
    if not tags:
        return "其他技术行业"
    merged = " ".join(tags).lower()
    for group, keywords in INDUSTRY_GROUP_RULES:
        if any(keyword.lower() in merged for keyword in keywords):
            return group
    return "其他技术行业"


def normalize_posting_snapshot(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    normalized["title"] = normalize_job_title(normalized.get("title"))
    normalized["location"] = _normalize_location_value(normalized.get("location"))
    normalized["salary_range"] = _normalize_text(normalized.get("salary_range"))
    normalized["company_name"] = _normalize_text(normalized.get("company_name"))
    normalized["industry"] = normalize_industry_value(normalized.get("industry"))
    normalized["industry_group"] = derive_industry_group(normalized.get("industry"))
    normalized["company_size"] = _normalize_company_size_value(normalized.get("company_size"))
    normalized["ownership_type"] = _normalize_ownership_type_value(normalized.get("ownership_type"))
    normalized["description"] = _normalize_text(normalized.get("description"))
    normalized["company_intro"] = _normalize_text(normalized.get("company_intro"))
    normalized["job_code"] = _normalize_text(normalized.get("job_code"))
    return normalized


def job_profile_snapshot(profile: Any) -> dict[str, Any]:
    return {
        "title": getattr(profile, "title", ""),
        "summary": getattr(profile, "summary", ""),
        "description": getattr(profile, "summary", ""),
        "skills": getattr(profile, "skill_requirements", []) or [],
        "job_code": getattr(profile, "job_code", ""),
    }


def _pick_value(row: dict[str, Any], aliases: tuple[str, ...]) -> str:
    normalized_row = {str(key).strip(): value for key, value in row.items()}
    for alias in aliases:
        for key, value in normalized_row.items():
            if key == alias:
                return _normalize_text(value)
    lowered = {key.lower(): value for key, value in normalized_row.items()}
    for alias in aliases:
        if alias.lower() in lowered:
            return _normalize_text(lowered[alias.lower()])
    return ""


def _fallback_job_code(title: str, company_name: str, source_url: str) -> str:
    digest = hashlib.md5(f"{title}|{company_name}|{source_url}".encode("utf-8")).hexdigest()[:12].upper()
    return f"A13-{digest}"


def normalize_job_title(value: Any) -> str:
    title = _normalize_text(value)
    if not title:
        return ""
    original_title = title
    for pattern in TITLE_CLEAN_PATTERNS:
        title = re.sub(pattern, " ", title)
    for noise in TITLE_NOISE_KEYWORDS:
        title = title.replace(noise, " ")
    lowered = original_title.lower()
    for keywords, canonical in TITLE_CANONICAL_RULES:
        if all(keyword in lowered for keyword in keywords):
            return canonical

    title = re.sub(r"[|丨]+", " ", title)
    title = re.sub(r"\s+", " ", title).strip(" -+")
    return title


def normalize_job_dataset_row(row: dict[str, Any]) -> dict[str, Any] | None:
    normalized = {field: _pick_value(row, aliases) for field, aliases in JOB_FIELD_ALIASES.items()}
    if not any(normalized.values()):
        return None
    normalized = normalize_posting_snapshot(normalized)
    if not normalized["title"]:
        return None
    normalized["source_url"] = normalized["source_url"].replace("http://", "https://")
    normalized["job_code"] = normalized["job_code"] or _fallback_job_code(
        normalized["title"],
        normalized["company_name"],
        normalized["source_url"],
    )
    return normalized


def is_computer_related_job(row: dict[str, Any]) -> bool:
    normalized_title = normalize_job_title(row.get("title"))
    title = _collapse_text(normalized_title)
    industry = _collapse_text(row.get("industry"))
    description = _collapse_text(row.get("description"))
    company_intro = _collapse_text(row.get("company_intro"))
    merged_text = " ".join(filter(None, (title, industry, description, company_intro)))

    strong_role_hit = any(keyword in title for keyword in TARGET_ROLE_KEYWORDS)
    tech_exception = any(keyword in title for keyword in TECH_EXCEPTION_KEYWORDS)
    strict_exclude = any(keyword in title for keyword in STRICT_EXCLUDE_TITLE_KEYWORDS)

    positive_score = 0
    if any(keyword in industry for keyword in COMPUTER_INDUSTRY_KEYWORDS):
        positive_score += 2
    if any(keyword in title for keyword in COMPUTER_ROLE_KEYWORDS):
        positive_score += 2
    positive_score += min(sum(1 for keyword in COMPUTER_SKILL_KEYWORDS if keyword in merged_text), 3)

    support_like = any(keyword in title for keyword in ("技术支持", "售前", "售后", "解决方案"))
    negative_role = any(keyword in title for keyword in NON_COMPUTER_ROLE_KEYWORDS)

    if strict_exclude and not tech_exception:
        return False
    if negative_role and positive_score < 2:
        return False
    if support_like and positive_score < 3:
        return False
    if strong_role_hit:
        return positive_score >= 2
    return any(keyword in industry for keyword in COMPUTER_INDUSTRY_KEYWORDS) and positive_score >= 4


def filter_computer_related_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if is_computer_related_job(row)]


def is_campus_relevant_job(row: dict[str, Any]) -> bool:
    title = _collapse_text(normalize_job_title(row.get("title")))
    if not title:
        return False

    job_code = _normalize_text(row.get("job_code")).upper()
    source_url = _collapse_text(row.get("source_url"))
    if job_code.startswith("CP-") or source_url.startswith("careerpilot://curated-campus/"):
        return True

    if any(keyword in title for keyword in CAMPUS_EXCLUDE_TITLE_KEYWORDS):
        return False

    industry = _collapse_text(row.get("industry"))
    description = _collapse_text(row.get("description"))
    company_intro = _collapse_text(row.get("company_intro"))
    merged_text = " ".join(filter(None, (title, industry, description, company_intro)))

    if is_computer_related_job(row):
        return True
    if any(keyword in title for keyword in CAMPUS_ROLE_KEYWORDS):
        return True
    industry_hits = sum(1 for keyword in CAMPUS_INDUSTRY_KEYWORDS if keyword in industry)
    role_hits = sum(1 for keyword in CAMPUS_ROLE_KEYWORDS if keyword in merged_text)
    return industry_hits >= 1 and role_hits >= 1


def filter_campus_relevant_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if is_campus_relevant_job(row)]


def _template_skill_blob(row: dict[str, Any]) -> str:
    skills = row.get("skills")
    if isinstance(skills, list):
        return " ".join(_normalize_text(item) for item in skills)
    return _normalize_text(skills)


def is_student_facing_computer_job(row: dict[str, Any]) -> bool:
    title = _collapse_text(normalize_job_title(row.get("title")))
    if not title:
        return False
    if any(keyword in title for keyword in STUDENT_FACING_EXCLUDE_KEYWORDS):
        return False

    merged_text = " ".join(
        filter(
            None,
            (
                title,
                _collapse_text(row.get("summary")),
                _collapse_text(row.get("description")),
                _collapse_text(_template_skill_blob(row)),
            ),
        )
    )
    if any(keyword in title for keyword in STUDENT_FACING_INCLUDE_KEYWORDS):
        return True
    title_supports_fallback = any(keyword in title for keyword in ("工程师", "开发", "分析师"))
    return title_supports_fallback and sum(
        1 for keyword in STUDENT_FACING_SKILL_KEYWORDS if keyword in merged_text
    ) >= 3


def filter_student_facing_job_templates(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if is_student_facing_computer_job(row)]


def filter_student_facing_job_profiles(profiles: list[Any]) -> list[Any]:
    return [profile for profile in profiles if is_student_facing_computer_job(job_profile_snapshot(profile))]


def _load_csv_rows(path: Path) -> list[dict[str, Any]]:
    encodings = ("utf-8-sig", "utf-8", "gbk")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                return [dict(row) for row in csv.DictReader(file)]
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return []


def _load_xls_rows(path: Path) -> list[dict[str, Any]]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:
        raise RuntimeError("读取 .xls 数据集需要安装 xlrd") from exc

    workbook = xlrd.open_workbook(path.as_posix())
    sheet = workbook.sheet_by_index(0)
    headers = [_normalize_text(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(1, sheet.nrows):
        rows.append({headers[col]: sheet.cell_value(row_idx, col) for col in range(sheet.ncols)})
    return rows


def _load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("读取 .xlsx 数据集需要安装 openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [_normalize_text(item) for item in values[0]]
    rows: list[dict[str, Any]] = []
    for raw_row in values[1:]:
        rows.append({headers[idx]: raw_row[idx] for idx in range(len(headers))})
    return rows


def resolve_job_dataset_path(explicit_path: str | None = None) -> Path:
    settings = get_settings()
    if settings.job_dataset_prefer_curated_local:
        for candidate in CURATED_DATASET_CANDIDATE_NAMES:
            path = settings.data_dir / candidate
            if path.exists():
                return path
    if explicit_path:
        path = Path(explicit_path).expanduser()
        if path.exists():
            return path
    if settings.job_dataset_path:
        path = Path(settings.job_dataset_path).expanduser()
        if path.exists():
            return path
    for candidate in DATASET_CANDIDATE_NAMES:
        path = settings.data_dir / candidate
        if path.exists():
            return path
    return settings.data_dir / "sample_jobs.csv"


def _load_dataset_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw_rows = _load_csv_rows(path)
    elif suffix == ".xls":
        raw_rows = _load_xls_rows(path)
    elif suffix == ".xlsx":
        raw_rows = _load_xlsx_rows(path)
    else:
        raise ValueError(f"不支持的数据集格式: {path.suffix}")

    normalized_rows: list[dict[str, Any]] = []
    for row in raw_rows:
        normalized = normalize_job_dataset_row(row)
        if normalized:
            normalized_rows.append(normalized)
    return normalized_rows


@lru_cache
def load_job_profile_templates() -> list[dict]:
    settings = get_settings()
    return _load_json(settings.data_dir / "job_profile_templates.json")


def load_job_graph_seed() -> dict:
    settings = get_settings()
    return _load_json(settings.data_dir / "job_graph_seed.json")


@lru_cache
def load_job_postings_dataset(
    dataset_path: str | None = None,
    computer_related_only: bool | None = None,
) -> list[dict]:
    path = resolve_job_dataset_path(dataset_path)
    rows = _load_dataset_rows(path)
    settings = get_settings()
    if computer_related_only is not None:
        return filter_computer_related_rows(rows) if computer_related_only else rows

    should_filter = settings.job_dataset_filtering_enabled
    if should_filter:
        scope = settings.job_dataset_scope
        if scope == "all":
            return rows
        if scope == "computer":
            return filter_computer_related_rows(rows)
        return filter_campus_relevant_rows(rows)
    return rows


def get_job_dataset_metadata(dataset_path: str | None = None) -> dict[str, Any]:
    path = resolve_job_dataset_path(dataset_path)
    raw_rows = _load_dataset_rows(path)
    filtered_rows = load_job_postings_dataset(str(path))
    return {
        "path": str(path),
        "file_name": path.name,
        "row_count": len(filtered_rows),
        "raw_row_count": len(raw_rows),
        "filtered_row_count": len(filtered_rows),
        "excluded_row_count": max(len(raw_rows) - len(filtered_rows), 0),
        "computer_related_only": get_settings().job_dataset_scope == "computer",
        "scope": get_settings().job_dataset_scope,
        "source": "sample" if path.name == "sample_jobs.csv" else ("curated" if path.name in CURATED_DATASET_CANDIDATE_NAMES else "official"),
    }


@lru_cache
def load_sample_job_postings() -> list[dict]:
    return load_job_postings_dataset()


def find_best_template(title: str) -> dict:
    title_lower = title.lower()
    for template in load_job_profile_templates():
        if template["title"].lower() in title_lower or title_lower in template["title"].lower():
            return template
    for template in load_job_profile_templates():
        if any(keyword.lower() in title_lower for keyword in template["skills"][:2]):
            return template
    return load_job_profile_templates()[0]
